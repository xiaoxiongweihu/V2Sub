"""术语表管理：增删改查 + JSON 存储 + CSV/Excel 导入导出。

数据结构：list[dict] -> [{"source": "API", "target": "应用程序接口"}, ...]
同时维护一个 (source.lower() -> target) 映射用于翻译时查找。
"""
from __future__ import annotations

import json
import os
import csv
from typing import Iterable


class Glossary:
    def __init__(self, path: str = "data/glossary.json") -> None:
        self.path = path
        self.entries: list[dict[str, str]] = []
        self.load()

    # ---- 持久化 ----
    def load(self) -> None:
        if os.path.isfile(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    self.entries = [
                        {"source": str(e.get("source", "")).strip(),
                         "target": str(e.get("target", "")).strip()}
                        for e in data
                        if e.get("source")
                    ]
            except (json.JSONDecodeError, OSError):
                self.entries = []

    def save(self) -> None:
        os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
        try:
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump(self.entries, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    # ---- 增删改查 ----
    def add(self, source: str, target: str) -> bool:
        source, target = source.strip(), target.strip()
        if not source:
            return False
        # 去重（按 source，大小写不敏感）
        for e in self.entries:
            if e["source"].lower() == source.lower():
                e["target"] = target
                self.save()
                return True
        self.entries.append({"source": source, "target": target})
        self.save()
        return True

    def update(self, index: int, source: str, target: str) -> bool:
        if not (0 <= index < len(self.entries)):
            return False
        source, target = source.strip(), target.strip()
        if not source:
            return False
        self.entries[index] = {"source": source, "target": target}
        self.save()
        return True

    def remove(self, index: int) -> bool:
        if not (0 <= index < len(self.entries)):
            return False
        self.entries.pop(index)
        self.save()
        return True

    def clear(self) -> None:
        self.entries.clear()
        self.save()

    def search(self, keyword: str) -> list[tuple[int, dict[str, str]]]:
        kw = keyword.strip().lower()
        if not kw:
            return [(i, e) for i, e in enumerate(self.entries)]
        return [(i, e) for i, e in enumerate(self.entries)
                if kw in e["source"].lower() or kw in e["target"].lower()]

    def __len__(self) -> int:
        return len(self.entries)

    # ---- 翻译提示注入 ----
    def hint_string(self, max_items: int = 80) -> str:
        """生成注入翻译 prompt 的术语提示串。空表返回空串。"""
        if not self.entries:
            return ""
        items = self.entries[:max_items]
        return "\n".join(f"{e['source']}={e['target']}" for e in items)

    # ---- 导入导出 ----
    def export_csv(self, path: str) -> int:
        """导出到 CSV，返回写入条数。"""
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["source", "target"])
            for e in self.entries:
                w.writerow([e["source"], e["target"]])
        return len(self.entries)

    def export_excel(self, path: str) -> int:
        """导出到 Excel，返回写入条数。"""
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise RuntimeError("导出 Excel 需 `pip install openpyxl`") from e
        wb = Workbook()
        ws = wb.active
        ws.title = "Glossary"
        ws.append(["source", "target"])
        for e in self.entries:
            ws.append([e["source"], e["target"]])
        wb.save(path)
        return len(self.entries)

    def import_file(self, path: str, merge: bool = True) -> int:
        """从 CSV/Excel 导入，返回新增条数。"""
        ext = os.path.splitext(path)[1].lower()
        rows: list[tuple[str, str]] = []
        if ext == ".csv":
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                start = 1 if header and header[0].lower() == "source" else 0
                f.seek(0)
                if start:
                    next(f)
                for row in csv.reader(f):
                    if len(row) >= 2 and row[0].strip():
                        rows.append((row[0].strip(), row[1].strip()))
        elif ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError as e:
                raise RuntimeError("导入 Excel 需 `pip install openpyxl`") from e
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 and row and str(row[0]).lower() == "source":
                    continue
                if row and len(row) >= 2 and row[0]:
                    rows.append((str(row[0]).strip(), str(row[1]).strip()))
        else:
            raise RuntimeError(f"不支持的文件类型: {ext}（支持 .csv/.xlsx）")

        added = 0
        if not merge:
            self.entries.clear()
        for s, t in rows:
            if self.add(s, t):
                # add 在已存在时会覆盖且返回 True，统一计为成功
                added += 1
        return added
